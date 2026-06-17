# app.py
# ------------------------------------------------------------
# Streamlit-app til at overføre musikrapport-data til tv-appl-en.xlsx
# ------------------------------------------------------------
#
# Kør lokalt sådan her i terminalen:
#   pip install -r requirements.txt
#   streamlit run app.py
#
# Vigtigt:
# - Appen bruger openpyxl til at åbne skabelonen og skrive værdier i eksisterende celler.
# - Den indsætter IKKE nye rækker og ændrer IKKE kolonnebredder, layout eller formatering.
# - Hvis der er flere sange end skabelonen har tomme rækker til, udfyldes kun de rækker,
#   der allerede findes i skabelonen. Appen viser en advarsel i browseren.

import io
import re
import traceback
import unicodedata
import zipfile
import time as time_module
from copy import copy
from datetime import datetime, time, timedelta
from difflib import SequenceMatcher

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# musicbrainzngs er et valgfrit dependency, som bruges til MusicBrainz-validering.
# Vi importerer det i en try/except, så appen stadig kan køre, selv hvis pakken
# ikke er installeret korrekt endnu.
try:
    import musicbrainzngs
except Exception:
    musicbrainzngs = None


# ------------------------------------------------------------
# 1. STREAMLIT-SIDEOPSÆTNING
# ------------------------------------------------------------

st.set_page_config(
    page_title="Musikrapport → TV application",
    page_icon="🎵",
    layout="centered",
)

st.title("🎵 Musikrapport → TV application")
st.markdown(
    """
Upload først NCB-skabelonen **tv-appl-en.xlsx** og derefter musikrapport-exporten.
Vælg derefter **land** og **produktionstype**, så appen kan bruge de relevante
NCB-regler for rapporteringen.

Når kildefilen er uploadet, finder appen automatisk alle faner. Du kan enten vælge
én bestemt fane eller udtrække alle faner på én gang.

Hvis du vælger alle faner, kan du vælge mellem:
- **Samlet musikrapport**: én Excel-fil med alle tracks samlet i den samme Music content-tabel.
- **Én musikrapport pr. fane**: en ZIP-fil med én separat Excel-rapport pr. kildefane.

Appen kan også automatisk samle gentagne **bumpers/vignetter** — fx norske bumpers —
så én gentaget bumper bliver skrevet som ét samlet linjeelement med “x antal”.
"""
)


# ------------------------------------------------------------
# 2. SMÅ HJÆLPEFUNKTIONER
# ------------------------------------------------------------

def is_empty(value) -> bool:
    """
    Tjekker om en celleværdi er tom.

    Excel-filer kan indeholde None, tomme strenge eller strenge med kun mellemrum.
    Denne funktion gør resten af koden mere robust.
    """
    return value is None or (isinstance(value, str) and value.strip() == "")


def safe_text(value) -> str:
    """
    Konverterer en celleværdi til tekst på en sikker måde.

    Hvis værdien er None, returneres en tom streng.
    Hvis værdien er et decimaltal som 39.0, returneres "39" i stedet for "39.0".
    """
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def normalize_text(value) -> str:
    """
    Normaliserer tekst, så vi lettere kan sammenligne labels og kolonnenavne.

    Eksempel:
    "Production Company:" bliver til "production company".
    """
    text = safe_text(value).lower().strip()
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    text = text.rstrip(":").strip()
    return text


def looks_like_label(value) -> bool:
    """
    Forsøger at vurdere, om en celle ligner et label i kildefilen.

    Mange metadatafelter i musikrapporten står som:
    "Production Title:"
    Derfor stopper vi med at lede mod højre, hvis vi rammer et nyt label.
    """
    return isinstance(value, str) and value.strip().endswith(":")


def unique_keep_order(values):
    """
    Fjerner dubletter, men bevarer rækkefølgen.

    Det er vigtigt ved komponister/forfattere:
    ["Smith", "Jones", "Smith"] bliver til ["Smith", "Jones"].
    """
    result = []
    seen = set()

    for value in values:
        text = safe_text(value)
        if not text:
            continue

        key = text.casefold()
        if key not in seen:
            seen.add(key)
            result.append(text)

    return result


# ------------------------------------------------------------
# 3. FUNKTIONER TIL AT FINDE METADATA I KILDEFILEN
# ------------------------------------------------------------

def find_value_next_to_label(ws, label_candidates, max_offset=6):
    """
    Finder værdien ud for et bestemt label i en Excel-fane.

    Eksempel i kildefilen:
    A12 = "Network/Station:"
    B12 = "TV 2"

    Funktionen finder label-cellen og returnerer den første ikke-tomme celle til højre.
    """
    labels = {normalize_text(label) for label in label_candidates}

    for row in ws.iter_rows():
        for cell in row:
            if normalize_text(cell.value) in labels:
                # Kig mod højre i samme række
                for offset in range(1, max_offset + 1):
                    target = ws.cell(row=cell.row, column=cell.column + offset)

                    # Hvis vi rammer et nyt label, stopper vi.
                    # Så undgår vi at komme til at læse et andet felt.
                    if looks_like_label(target.value):
                        return None

                    if not is_empty(target.value):
                        return target.value

                return None

    return None


def find_cell_right_of_label(ws, label_candidates):
    """
    Finder cellen lige til højre for et label i skabelonen.

    Eksempel i skabelonen:
    A3 = "Production company"
    B3 = tom celle, som skal udfyldes

    Funktionen returnerer selve celle-objektet, så vi kan skrive en værdi i den.
    """
    labels = {normalize_text(label) for label in label_candidates}

    for row in ws.iter_rows():
        for cell in row:
            if normalize_text(cell.value) in labels:
                return ws.cell(row=cell.row, column=cell.column + 1)

    return None


def extract_metadata(source_ws) -> dict:
    """
    Udtrækker metadata fra den valgte kildefane.

    Mapping:
    - Production company = Production Company:
    - Production title   = Production Title:
    - Episode number     = Episode Number:
    - Broadcaster        = Network/Station:
    """
    return {
        "Production company": find_value_next_to_label(
            source_ws,
            ["Production Company:"]
        ),
        "Production title": find_value_next_to_label(
            source_ws,
            ["Production Title:"]
        ),
        "Episode number": find_value_next_to_label(
            source_ws,
            ["Episode Number:"]
        ),
        "Broadcaster": find_value_next_to_label(
            source_ws,
            ["Network/Station:"]
        ),
    }


# ------------------------------------------------------------
# 4. FUNKTIONER TIL AT FINDE HEADER-RÆKKER OG KOLONNER
# ------------------------------------------------------------

def find_header_row(ws, required_headers):
    """
    Finder den række, hvor bestemte kolonneoverskrifter findes.

    I kildefilen leder vi fx efter:
    "Seq #" og "Music Title".

    I skabelonen leder vi fx efter:
    "Song title", "Artist", "Min", "Sec", "Track type".
    """
    required = {normalize_text(header) for header in required_headers}

    for row in ws.iter_rows():
        header_map = {}

        for cell in row:
            if not is_empty(cell.value):
                header_map[normalize_text(cell.value)] = cell.column

        if required.issubset(set(header_map.keys())):
            return row[0].row, header_map

    raise ValueError(
        f"Kunne ikke finde header-rækken med disse kolonner: {required_headers}"
    )


def get_col(header_map, candidates, required=True):
    """
    Finder en kolonne ud fra én eller flere mulige kolonnenavne.

    Eksempel:
    get_col(header_map, ["Music Performer", "Associated performer"])
    """
    for candidate in candidates:
        key = normalize_text(candidate)
        if key in header_map:
            return header_map[key]

    if required:
        raise ValueError(f"Mangler kolonne i filen. Prøvede: {candidates}")

    return None


def get_col_contains(header_map, candidates, required=True):
    """
    Finder en kolonne, hvor kolonnenavnet indeholder en bestemt tekst.

    Det bruges især til skabelonen, hvor headeren fx hedder:
    "Composers (Last name). Separate with /"

    Her vil vi gerne finde den bare ved at søge på "Composers".
    """
    for candidate in candidates:
        needle = normalize_text(candidate)

        for header_name, col_index in header_map.items():
            if needle == header_name or needle in header_name:
                return col_index

    if required:
        raise ValueError(f"Mangler kolonne i skabelonen. Prøvede: {candidates}")

    return None


# ------------------------------------------------------------
# 5. NAVNE-RENSNING: KOMPONISTER OG FORFATTERE
# ------------------------------------------------------------

def clean_party_name(name) -> str:
    """
    Rydder op i navne fra "Name of Music Interested Party".

    Fjerner fx:
    - [PRS]
    - [BMI]
    - (PRS)
    - (BMI)

    Eksempel:
    "John Williams [BMI]" bliver til "John Williams".
    """
    text = safe_text(name)

    # Fjern alt i kantede parenteser: [PRS]
    text = re.sub(r"\[[^\]]*\]", "", text)

    # Fjern alt i almindelige parenteser: (BMI)
    text = re.sub(r"\([^)]*\)", "", text)

    # Ryd op i mellemrum og tegn
    text = text.replace(";", " ")
    text = re.sub(r"\s+", " ", text).strip(" ,;/")

    return text


def split_party_names(raw_name):
    """
    Splitter navne, hvis flere navne mod forventning står i samme celle.

    Normalt står der ét navn per række, men denne funktion gør appen mere robust.
    Vi splitter forsigtigt på / og ;.
    """
    cleaned = clean_party_name(raw_name)

    if not cleaned:
        return []

    parts = re.split(r"\s*/\s*|\s*;\s*", cleaned)
    return [part.strip() for part in parts if part.strip()]


def extract_last_name(name) -> str:
    """
    Udtrækker efternavnet fra et navn.

    Eksempler:
    - "A. Z. Børresen" bliver til "Børresen"
    - "John Williams" bliver til "Williams"
    - "Sting" bliver til "Sting"
    - "Williams, John" bliver til "Williams"
    """
    text = clean_party_name(name)

    if not text:
        return ""

    # Hvis navnet står som "Efternavn, Fornavn", tager vi delen før kommaet.
    if "," in text:
        return text.split(",", 1)[0].strip()

    parts = text.split()

    # Ét ord: fx "Sting" eller "EJAE"
    if len(parts) == 1:
        return parts[0]

    # Håndter navne med suffix, fx "Tom Jesso Jr."
    suffixes = {"jr", "jr.", "sr", "sr.", "ii", "iii", "iv", "v"}
    last_token_normalized = parts[-1].lower().strip(".")

    if last_token_normalized in {suffix.strip(".") for suffix in suffixes} and len(parts) >= 2:
        return f"{parts[-2]} {parts[-1]}"

    # Standard: sidste ord er efternavnet
    return parts[-1]


# ------------------------------------------------------------
# 6. TIDS-PARSING: MUSIC DURATION → MIN/SEC
# ------------------------------------------------------------

def parse_duration_to_min_sec(value):
    """
    Konverterer "Music Duration" til minutter og sekunder.

    Funktionen forsøger at håndtere flere formater:
    - "00:02:15"
    - "02:15"
    - "0.02:26"
    - Excel time-værdier som decimaltal
    - datetime.time
    - timedelta
    """
    if is_empty(value):
        return None, None

    total_seconds = None

    # Hvis Excel/openpyxl giver os en timedelta
    if isinstance(value, timedelta):
        total_seconds = int(round(value.total_seconds()))

    # Hvis værdien er en datetime
    elif isinstance(value, datetime):
        total_seconds = value.hour * 3600 + value.minute * 60 + value.second

    # Hvis værdien er et time-objekt
    elif isinstance(value, time):
        total_seconds = value.hour * 3600 + value.minute * 60 + value.second

    # Hvis værdien er et tal
    elif isinstance(value, (int, float)):
        # Excel gemmer nogle gange tid som en brøkdel af et døgn.
        # 0.5 = 12 timer.
        if value < 1:
            total_seconds = int(round(value * 24 * 60 * 60))
        else:
            # Hvis tallet er større end 1, antager vi, at det er sekunder.
            total_seconds = int(round(value))

    # Hvis værdien er tekst
    else:
        text = safe_text(value).strip()

        # Nogle exports kan skrive 0.02:26 i stedet for 00:02:26.
        # Vi gør det læsbart ved at skifte punktum til kolon, hvis der også er kolon.
        text = text.replace(",", ":")
        if "." in text and ":" in text:
            text = text.replace(".", ":")

        # Fjern mellemrum og udtræk tal
        text = re.sub(r"\s+", "", text)
        numbers = re.findall(r"\d+", text)

        if len(numbers) >= 3:
            # Brug de sidste tre tal som timer, minutter, sekunder
            hours = int(numbers[-3])
            minutes = int(numbers[-2])
            seconds = int(numbers[-1])
            total_seconds = hours * 3600 + minutes * 60 + seconds

        elif len(numbers) == 2:
            # MM:SS
            minutes = int(numbers[0])
            seconds = int(numbers[1])
            total_seconds = minutes * 60 + seconds

        elif len(numbers) == 1:
            # Kun ét tal: antag sekunder
            total_seconds = int(numbers[0])

    if total_seconds is None:
        return None, None

    if total_seconds < 0:
        total_seconds = 0

    minutes = total_seconds // 60
    seconds = total_seconds % 60

    return int(minutes), int(seconds)


# ------------------------------------------------------------
# 7. TRACK TYPE-MAPPING
# ------------------------------------------------------------

TRACK_TYPE_MAP = {
    "pre-existing": "Existing music",
    "pre existing": "Existing music",
    "preexisting": "Existing music",
    "specially commissioned": "Commissioned music",
    "special commissioned": "Commissioned music",
    "commissioned": "Commissioned music",
    "library/production": "Library music",
    "library production": "Library music",
    "library": "Library music",
    "production music": "Library music",
}


def translate_track_type(value) -> str:
    """
    Oversætter "Music Source" fra kildefilen til skabelonens Track type.

    Krævet mapping:
    - Pre-existing → Existing music
    - Specially Commissioned → Commissioned music
    - Library/Production → Library music
    """
    text = normalize_text(value)

    if not text:
        return ""

    if text in TRACK_TYPE_MAP:
        return TRACK_TYPE_MAP[text]

    # Fallback: hvis teksten indeholder et kendt mønster
    for key, mapped_value in TRACK_TYPE_MAP.items():
        if key in text:
            return mapped_value

    # Hvis værdien er ukendt, returnerer vi den oprindelige tekst.
    # Så kan brugeren se, at der var noget, der ikke matchede standardmappingen.
    return safe_text(value)



# ------------------------------------------------------------
# 8. LANDE, PRODUKTIONSTYPER OG SMART RAPPORT-REGLER
# ------------------------------------------------------------

COUNTRY_CONFIG = {
    "Danmark": {
        "territory": "Denmark / Nordic region",
        "production_types": [
            "TV-produktion",
            "Live on tape",
            "Fiktionsserie (Norden)",
            "Fiktionsserie (udenfor Norden)",
            "Koncert",
            "Vignet og bumper",
            "Trailer",
            "TV-reklame",
            "Reklame & trailers",
            "Spillefilm",
            "Kortfilm & dokumentar - biograf",
            "Kortfilm & dokumentar - festival/TV/online",
            "Fysiske kopier - AV-produktioner",
        ],
        "rule_note": "Danmark: Vignet og bumper ligger i samme kategori. Bestillingsmusik er angivet pr. episode; eksisterende musik forhandles direkte.",
    },
    "Finland": {
        "territory": "Finland / Nordic region",
        "production_types": [
            "Muut TV-ohjelmat / andre TV-programmer",
            "Live on tape",
            "TV-drama (Norden)",
            "TV-drama (udenfor Norden)",
            "Koncertoptagelse",
            "Tunnus- ja bumpermusiikki / signature og bumper",
            "Reklamer, trailers og promos",
            "Feature film",
            "Short film - theatrical",
            "Short film - festival/TV/online",
            "Documentary - theatrical",
            "Documentary - TV/festival/online",
            "Fysiske kopier - audiovisuelle optagelser",
        ],
        "rule_note": "Finland: Signature-/bumpermusik er samlet i én kategori. Eksisterende musik forhandles direkte; bestillingsmusik er angivet pr. musiksekund.",
    },
    "Island": {
        "territory": "Iceland / Nordic region",
        "production_types": [
            "TV production",
            "Live on tape",
            "TV drama (Norden)",
            "TV drama (udenfor Norden)",
            "Concert recording",
            "Signature and bumper",
            "TV trailer",
            "TV commercial / TV promo",
            "Commercial / promo & trailer",
            "Feature film",
            "Short film & documentary - theatrical",
            "Short film & documentary - non-theatrical",
            "Physical copies - audiovisual productions",
        ],
        "rule_note": "Island: Signature and bumper er én kategori og forhandles direkte med rettighedshaver.",
    },
    "Norge": {
        "territory": "Norway / Nordic region",
        "production_types": [
            "TV-produksjon",
            "Live on tape",
            "TV-drama (Norden)",
            "TV-drama (udenfor Norden)",
            "Konsert",
            "Vignett",
            "Bumper",
            "Trailer & TV-reklame",
            "Spillefilm",
            "Kortfilm - kinodistribusjon",
            "Kortfilm - TV/festival/annen visning",
            "Dokumentar - kinodistribusjon",
            "Dokumentar - TV/festival/annen visning",
            "Fysiske kopier - AV-produksjoner",
        ],
        "rule_note": "Norge: Vignett og Bumper er separate kategorier. Appen kan derfor automatisk samle gentagne bumpers/vignetter og markere dem med x antal.",
    },
    "Sverige": {
        "territory": "Sweden / Nordic region",
        "production_types": [
            "TV-produktion",
            "Live on tape",
            "TV-drama (Norden)",
            "TV-drama (udenfor Norden)",
            "Koncerter",
            "Vinjett & bumper",
            "TV-trailer",
            "TV-reklam",
            "Reklam & trailer",
            "Spelfilm",
            "Kortfilm & dokumentär - biograf",
            "Kortfilm & dokumentär - uden biografvisning",
            "Fysiske kopier - AV-produktioner",
        ],
        "rule_note": "Sverige: Vinjett & bumper er én kategori og forhandles direkte med rettighedshaver.",
    },
}


def get_country_names():
    """Returnerer landene i den rækkefølge, de skal vises i brugerfladen."""
    return list(COUNTRY_CONFIG.keys())


def get_production_types_for_country(country: str) -> list[str]:
    """Finder produktionstyperne for det valgte land."""
    return COUNTRY_CONFIG.get(country, COUNTRY_CONFIG["Danmark"])["production_types"]


def seconds_from_song(song: dict) -> int:
    """
    Regner en sangs Min/Sec om til totale sekunder.

    Det gør det muligt at samle 12 ens bumpers til én linje med samlet varighed.
    """
    minutes = song.get("Min") or 0
    seconds = song.get("Sec") or 0

    try:
        return int(minutes) * 60 + int(seconds)
    except Exception:
        return 0


def min_sec_from_seconds(total_seconds: int):
    """Konverterer totale sekunder tilbage til Min/Sec-kolonnerne."""
    total_seconds = max(0, int(total_seconds or 0))
    return total_seconds // 60, total_seconds % 60


def clean_title_for_grouping(title: str) -> str:
    """
    Rydder titel op, så små forskelle ikke ødelægger samlingen.

    Eksempel:
    "03 BUMPER TO COMMERCIAL BREAK" og "03 Bumper To Commercial Break"
    bliver behandlet som samme titel.
    """
    text = safe_text(title)
    text = re.sub(r"\s+x\s*\d+\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def detect_report_category(song: dict, selected_country: str, production_type: str) -> str:
    """
    Gætter på om et track er almindelig musik, bumper eller vignette.

    Vi bruger både den valgte produktionstype, sangtitel og Music Usage.
    Det er især nyttigt i norske rapporter, hvor Vignett og Bumper har hver sin regel.
    """
    title = normalize_text(song.get("Song title"))
    usage = normalize_text(song.get("Music Usage"))
    production = normalize_text(production_type)

    # Hvis brugeren direkte har valgt Bumper/Vignett som produktionstype,
    # behandler vi alle tracks som den type.
    if "bumper" in production and not any(word in production for word in ["vign", "vinjett", "signature", "tunnus"]):
        return "Bumper"

    if any(word in production for word in ["vignett", "vignet", "vinjett", "signature", "tunnus"] ) and "bumper" not in production:
        return "Vignette"

    # Automatisk genkendelse ud fra titel.
    bumper_words = [
        "bumper",
        "bumpers",
        "commercial break",
        "reklamepause",
        "reklamepauser",
        "reklampaus",
    ]
    if any(word in title for word in bumper_words):
        return "Bumper"

    vignette_words = [
        "vignett",
        "vignet",
        "vinjett",
        "signature",
        "theme",
        "kjenningsmelodi",
        "kendingsmelodi",
        "tunnus",
        "opening theme",
        "closing theme",
    ]
    if any(word in title for word in vignette_words) or usage in ["opening theme", "closing theme"]:
        return "Vignette"

    return "Music"


def should_aggregate_song(song: dict, selected_country: str, production_type: str, auto_group_special: bool) -> tuple[bool, str]:
    """
    Beslutter om et track skal samles med andre ens tracks.

    Appens hovedregel:
    - Bumpers/vignetter kan samles, fordi de ofte gentages mange gange.
    - Almindelig musik samles ikke automatisk, da hvert cue normalt skal stå for sig.
    - Norge får særlig opmærksomhed, fordi Vignett og Bumper er separate regler.
    """
    if not auto_group_special:
        return False, "Music"

    category = detect_report_category(song, selected_country, production_type)
    production = normalize_text(production_type)
    country = normalize_text(selected_country)

    # Hvis produktionstypen selv er en bumper-/vignette-kategori, samler vi.
    if any(word in production for word in ["bumper", "vign", "vinjett", "signature", "tunnus"]):
        if category == "Music":
            category = "Vignette/Bumper"
        return True, category

    # Norsk special: Hvis en almindelig TV-rapport indeholder bumpers/vignetter,
    # samles de automatisk, så uploadarbejdet bliver mindre manuelt.
    if country == "norge" and category in ["Bumper", "Vignette"]:
        return True, category

    return False, category


def format_aggregated_title(original_title: str, category: str, count: int) -> str:
    """
    Laver den titel, der skrives i rapporten for en samlet bumper/vignette.

    Eksempel:
    "03 BUMPER TO COMMERCIAL BREAK x 12"
    "Bumper - THE VOICE OF... x 8"
    """
    title = clean_title_for_grouping(original_title)
    lower_title = normalize_text(title)

    if category == "Bumper" and "bumper" not in lower_title:
        title = f"Bumper - {title}"
    elif category == "Vignette" and not any(word in lower_title for word in ["vignett", "vignet", "vinjett", "signature", "theme", "tunnus"]):
        title = f"Vignette - {title}"

    if count > 1:
        title = f"{title} x {count}"

    return title


def aggregate_repeated_special_music(songs: list[dict], selected_country: str, production_type: str, auto_group_special: bool = True):
    """
    Samler gentagne bumpers/vignetter efter land og produktionstype.

    Gruppens nøgle er titel + artist + komponister + writers + track type + kategori.
    Varigheden lægges sammen, og titlen får "x antal".
    """
    output = []
    groups = {}
    order = []
    aggregated_original_rows = 0

    for song in songs:
        should_group, category = should_aggregate_song(
            song=song,
            selected_country=selected_country,
            production_type=production_type,
            auto_group_special=auto_group_special,
        )

        if not should_group:
            output.append(song)
            continue

        key = (
            category,
            normalize_text(clean_title_for_grouping(song.get("Song title"))),
            normalize_text(song.get("Artist")),
            normalize_text(song.get("Composers")),
            normalize_text(song.get("Writers")),
            normalize_text(song.get("Track type")),
        )

        if key not in groups:
            groups[key] = {
                "base_song": dict(song),
                "category": category,
                "count": 0,
                "seconds": 0,
            }
            order.append(key)

        groups[key]["count"] += 1
        groups[key]["seconds"] += seconds_from_song(song)
        aggregated_original_rows += 1

    # De samlede linjer lægges efter de almindelige tracks.
    # Det gør dem lette at finde og indtaste i næste system.
    for key in order:
        group = groups[key]
        song = dict(group["base_song"])
        minutes, seconds = min_sec_from_seconds(group["seconds"])

        song["Song title"] = format_aggregated_title(
            original_title=song.get("Song title"),
            category=group["category"],
            count=group["count"],
        )
        song["Min"] = minutes
        song["Sec"] = seconds
        song["Aggregated count"] = group["count"]
        song["Aggregated category"] = group["category"]

        output.append(song)

    collapsed_lines = max(0, aggregated_original_rows - len(groups))
    warnings = []

    if collapsed_lines > 0:
        warnings.append(
            f"{aggregated_original_rows} bumper-/vignette-linjer blev samlet til {len(groups)} linjer "
            f"efter reglerne for {selected_country} / {production_type}."
        )

    return output, {
        "aggregated_original_rows": aggregated_original_rows,
        "aggregated_groups": len(groups),
        "collapsed_lines": collapsed_lines,
        "warnings": warnings,
    }


def apply_country_production_rules(songs: list[dict], selected_country: str, production_type: str, auto_group_special: bool = True):
    """
    Samlet regelmotor.

    Lige nu gør den især én vigtig ting:
    - Den samler gentagne bumpers/vignetter, når land/produktionstype peger på det.

    Funktionen returnerer både de bearbejdede sange og en lille summary til UI'et.
    """
    processed_songs, rule_summary = aggregate_repeated_special_music(
        songs=songs,
        selected_country=selected_country,
        production_type=production_type,
        auto_group_special=auto_group_special,
    )

    rule_summary["country"] = selected_country
    rule_summary["production_type"] = production_type
    rule_summary["auto_group_special"] = auto_group_special

    return processed_songs, rule_summary


def write_report_context_to_template(template_ws, selected_country: str, production_type: str):
    """
    Skriver land og produktionstype ind i skabelonen, hvis felterne findes.

    Skabelonen har fx "Production category" og "Territory".
    Hvis en fremtidig skabelon ikke har felterne, springer funktionen bare over.
    """
    country_config = COUNTRY_CONFIG.get(selected_country, {})
    territory = country_config.get("territory", selected_country)

    write_optional_metadata_value(template_ws, "Production category", production_type)
    write_optional_metadata_value(template_ws, "Territory", territory)


# ------------------------------------------------------------
# 8. UDTRÆK MUSIKDATA FRA KILDEFILEN
# ------------------------------------------------------------

def extract_music_rows(source_ws) -> list[dict]:
    """
    Udtrækker alle sange fra den valgte kildefane.

    Logik:
    - Find header-rækken med "Seq #" og "Music Title"
    - Gruppér efter Seq #
    - Linjer under en sang uden Seq # tilhører seneste Seq #
    - Tomme rækker ignoreres
    """
    header_row, header_map = find_header_row(
        source_ws,
        required_headers=["Seq #", "Music Title"]
    )

    # Find de relevante kolonner i kildefilen
    col_seq = get_col(header_map, ["Seq #"])
    col_title = get_col(header_map, ["Music Title"])
    col_duration = get_col(header_map, ["Music Duration"])
    col_source = get_col(header_map, ["Music Source"])
    col_usage = get_col(header_map, ["Music Usage"], required=False)

    # Ikke alle fremtidige exports er nødvendigvis helt ens,
    # så disse tre gøres ikke hårdt påkrævede.
    col_performer = get_col(header_map, ["Music Performer"], required=False)
    col_role = get_col(header_map, ["Music Interested Party Role"], required=False)
    col_name = get_col(header_map, ["Name of Music Interested Party"], required=False)

    groups = []
    current_group = None

    # Gå igennem alle rækker under headeren
    for row_number in range(header_row + 1, source_ws.max_row + 1):
        row_values = [
            source_ws.cell(row=row_number, column=col).value
            for col in range(1, source_ws.max_column + 1)
        ]

        # Spring helt tomme rækker over
        if all(is_empty(value) for value in row_values):
            continue

        seq_value = source_ws.cell(row=row_number, column=col_seq).value

        row_data = {
            "seq": seq_value,
            "title": source_ws.cell(row=row_number, column=col_title).value,
            "duration": source_ws.cell(row=row_number, column=col_duration).value,
            "source": source_ws.cell(row=row_number, column=col_source).value,
            "usage": source_ws.cell(row=row_number, column=col_usage).value if col_usage else None,
            "performer": source_ws.cell(row=row_number, column=col_performer).value if col_performer else None,
            "role": source_ws.cell(row=row_number, column=col_role).value if col_role else None,
            "name": source_ws.cell(row=row_number, column=col_name).value if col_name else None,
        }

        # Hvis rækken har Seq #, starter en ny sang
        if not is_empty(seq_value):
            current_group = {
                "seq": seq_value,
                "rows": [row_data],
            }
            groups.append(current_group)

        # Hvis rækken ikke har Seq #, hører den til seneste sang
        elif current_group is not None:
            current_group["rows"].append(row_data)

    songs = []

    for group in groups:
        rows = group["rows"]

        # Find første ikke-tomme titel, performer, duration og source i gruppen
        title = next(
            (safe_text(row["title"]) for row in rows if not is_empty(row["title"])),
            ""
        )

        performer = next(
            (safe_text(row["performer"]) for row in rows if not is_empty(row["performer"])),
            ""
        )

        duration = next(
            (row["duration"] for row in rows if not is_empty(row["duration"])),
            None
        )

        usage = next(
            (safe_text(row["usage"]) for row in rows if not is_empty(row.get("usage"))),
            ""
        )

        source = next(
            (row["source"] for row in rows if not is_empty(row["source"])),
            None
        )

        # Hvis Music Performer ikke er udfyldt, prøver vi Associated performer-linjer
        if not performer:
            associated_performers = []

            for row in rows:
                role = normalize_text(row["role"])

                if "associated performer" in role and not is_empty(row["name"]):
                    associated_performers.append(clean_party_name(row["name"]))

            performer = "/".join(unique_keep_order(associated_performers))

        composers = []
        writers = []

        for row in rows:
            role = normalize_text(row["role"])
            raw_name = row["name"]

            if is_empty(role) or is_empty(raw_name):
                continue

            # Split i tilfælde af flere navne i samme celle
            for one_name in split_party_names(raw_name):
                last_name = extract_last_name(one_name)

                if not last_name:
                    continue

                # Composer og Composer/Author ryger i komponistfeltet
                if "composer" in role:
                    composers.append(last_name)

                # Author, Writer, Lyricist og Composer/Author ryger i writerfeltet
                if "author" in role or "writer" in role or "lyricist" in role:
                    writers.append(last_name)

        minutes, seconds = parse_duration_to_min_sec(duration)

        songs.append(
            {
                "seq": group["seq"],
                "Song title": title,
                "Artist": performer,
                "Composers": "/".join(unique_keep_order(composers)),
                "Writers": "/".join(unique_keep_order(writers)),
                "Min": minutes,
                "Sec": seconds,
                "Track type": translate_track_type(source),
                "Music Usage": usage,
            }
        )

    return songs


# ------------------------------------------------------------
# 8B. MUSICBRAINZ-VALIDERING AF KOMPONISTER
# ------------------------------------------------------------

MUSICBRAINZ_WRITER_RELATION_WORDS = [
    "composer",
    "writer",
    "author",
    "lyricist",
    "librettist",
]


def normalize_for_match(value: str) -> str:
    """
    Gør navne/titler lettere at sammenligne.

    Eksempel:
    "Björk Guðmundsdóttir" og "Bjork Gudmundsdottir" bliver mere ens,
    fordi accenter og specialtegn fjernes.
    """
    text = safe_text(value).casefold().strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9æøå ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def split_slash_names(value: str) -> list[str]:
    """
    Splitter komponist-/writer-felter fra rapporten.

    Skabelonen bruger normalt slash mellem efternavne:
    "Lennon/McCartney" → ["Lennon", "McCartney"]
    """
    text = safe_text(value)

    if not text:
        return []

    return [part.strip() for part in re.split(r"\s*/\s*|\s*;\s*", text) if part.strip()]


def get_report_writer_names(song: dict) -> list[str]:
    """
    Henter de navne, som allerede står i rapporten.

    Vi kombinerer både Composers og Writers, fordi MusicBrainz kan bruge
    forskellige relationstyper afhængigt af værket.
    """
    names = []
    names.extend(split_slash_names(song.get("Composers", "")))
    names.extend(split_slash_names(song.get("Writers", "")))
    return unique_keep_order(names)


def is_existing_music(song: dict) -> bool:
    """
    Tjekker om et track er Existing music.

    MusicBrainz-valideringen skal kun køres på eksisterende musik, ikke på
    library music eller commissioned music.
    """
    return normalize_text(song.get("Track type")) == "existing music"


def setup_musicbrainz_client():
    """
    Klargør MusicBrainz-klienten.

    MusicBrainz kræver, at API-kald identificerer applikationen med en user-agent.
    Hvis musicbrainzngs ikke er installeret, returnerer funktionen en pæn fejltekst
    i stedet for at crashe appen.
    """
    if musicbrainzngs is None:
        return False, "Pakken musicbrainzngs er ikke installeret. Tjek requirements.txt."

    try:
        musicbrainzngs.set_useragent(
            "musikrapport-streamlit-app",
            "1.0",
            "https://streamlit.app",
        )
        return True, ""
    except Exception as error:
        return False, f"Kunne ikke klargøre MusicBrainz-klienten: {error}"


def extract_artist_name_from_relation(relation: dict) -> str:
    """
    Finder kunstner-/personnavn i en MusicBrainz-relation.

    MusicBrainz-data kan variere en smule, så vi prøver flere mulige nøgler.
    """
    artist = relation.get("artist") or {}

    if isinstance(artist, dict):
        return safe_text(
            artist.get("name")
            or artist.get("sort-name")
            or artist.get("artist-credit-phrase")
        )

    return safe_text(relation.get("name") or relation.get("target"))


def extract_musicbrainz_writer_names(work: dict) -> list[str]:
    """
    Udtrækker komponist-/forfatternavne fra et MusicBrainz-work.

    Vi leder efter artist-relationer med typer som composer, writer, author,
    lyricist eller librettist.
    """
    names = []

    relation_lists = []

    # Typisk placering i musicbrainzngs for work → artist-relationer.
    if isinstance(work.get("artist-relation-list"), list):
        relation_lists.append(work.get("artist-relation-list"))

    # Fallback, hvis API'et returnerer en mere generisk relation-list.
    if isinstance(work.get("relation-list"), list):
        relation_lists.append(work.get("relation-list"))

    for relation_list in relation_lists:
        for relation in relation_list:
            relation_type = normalize_text(relation.get("type"))
            target_type = normalize_text(relation.get("target-type"))

            # Vi vil kun bruge person-/artist-relationer, ikke fx URL'er.
            if target_type and target_type != "artist":
                continue

            if any(word in relation_type for word in MUSICBRAINZ_WRITER_RELATION_WORDS):
                name = extract_artist_name_from_relation(relation)
                if name:
                    names.append(name)

    return unique_keep_order(names)


def similarity(a: str, b: str) -> float:
    """
    Lille fuzzy matching-funktion, så små variationer ikke ødelægger et match.
    """
    return SequenceMatcher(None, normalize_for_match(a), normalize_for_match(b)).ratio()


@st.cache_data(ttl=24 * 60 * 60, show_spinner=False)
def musicbrainz_lookup_work_writers(title: str, artist: str, max_candidates: int = 3) -> dict:
    """
    Slår et værk op i MusicBrainz og returnerer mulige writer/composer-navne.

    Funktionen er cachet i 24 timer, så samme sang ikke slår API'et igen og igen,
    hver gang Streamlit genindlæser appen.
    """
    ok, setup_error = setup_musicbrainz_client()
    if not ok:
        return {
            "status": "error",
            "message": setup_error,
            "candidates": [],
        }

    title = safe_text(title)
    artist = safe_text(artist)

    if not title:
        return {
            "status": "not_found",
            "message": "Ingen titel at søge på.",
            "candidates": [],
        }

    candidates = []
    seen_work_ids = set()

    # Vi prøver først med både work/titel og artist. Hvis det ikke giver noget,
    # prøver vi kun med titel. Det gør funktionen mere robust for obskure tracks.
    search_attempts = []
    if artist:
        search_attempts.append({"work": title, "artist": artist})
    search_attempts.append({"work": title})

    try:
        for fields in search_attempts:
            try:
                result = musicbrainzngs.search_works(
                    limit=max_candidates,
                    strict=False,
                    **fields,
                )
                work_list = result.get("work-list", []) or []
            except Exception:
                # Hvis artist-søgningen fejler, prøver vi næste søgning.
                work_list = []

            for work_item in work_list[:max_candidates]:
                work_id = work_item.get("id")
                if not work_id or work_id in seen_work_ids:
                    continue

                seen_work_ids.add(work_id)

                try:
                    detail = musicbrainzngs.get_work_by_id(
                        work_id,
                        includes=["artist-rels"],
                    )
                    work = detail.get("work", {}) or {}
                except Exception:
                    work = work_item

                writer_names = extract_musicbrainz_writer_names(work)
                work_title = safe_text(work.get("title") or work_item.get("title"))

                candidates.append(
                    {
                        "id": work_id,
                        "title": work_title,
                        "score": safe_text(work_item.get("score")),
                        "writers": writer_names,
                    }
                )

            if candidates:
                break

        if not candidates:
            return {
                "status": "not_found",
                "message": "MusicBrainz fandt ikke et sikkert work-match.",
                "candidates": [],
            }

        return {
            "status": "ok",
            "message": "",
            "candidates": candidates,
        }

    except Exception as error:
        return {
            "status": "error",
            "message": str(error),
            "candidates": [],
        }


def compare_report_names_to_musicbrainz(report_names: list[str], mb_names: list[str]) -> dict:
    """
    Sammenligner rapportens komponister/writers med MusicBrainz-navne.

    Vi sammenligner primært efternavne, fordi skabelonen kun kræver efternavne.
    """
    report_last_names = []
    for name in report_names:
        last_name = extract_last_name(name)
        if last_name:
            report_last_names.append(last_name)

    mb_last_names = []
    for name in mb_names:
        last_name = extract_last_name(name)
        if last_name:
            mb_last_names.append(last_name)

    report_keys = [normalize_for_match(name) for name in unique_keep_order(report_last_names)]
    mb_keys = [normalize_for_match(name) for name in unique_keep_order(mb_last_names)]

    missing = []
    matched = []

    for original_name, key in zip(unique_keep_order(report_last_names), report_keys):
        # Først prøver vi direkte match på efternavn.
        direct_match = key in mb_keys

        # Derefter et lille fuzzy fallback, hvis der fx er en accent eller stavemåde,
        # der afviger en smule.
        fuzzy_match = any(similarity(key, mb_key) >= 0.86 for mb_key in mb_keys)

        if direct_match or fuzzy_match:
            matched.append(original_name)
        else:
            missing.append(original_name)

    return {
        "all_match": bool(report_keys) and not missing,
        "matched": matched,
        "missing": missing,
        "musicbrainz_last_names": unique_keep_order(mb_last_names),
        "musicbrainz_full_names": unique_keep_order(mb_names),
    }


def validate_one_song_with_musicbrainz(song: dict) -> dict:
    """
    Validerer én sang mod MusicBrainz.

    Returnerer samme song-dict, men med en ekstra nøgle:
    - Data Match
    """
    validated_song = dict(song)

    if not is_existing_music(validated_song):
        validated_song["Data Match"] = "— Ikke Existing music"
        return validated_song

    report_names = get_report_writer_names(validated_song)

    if not report_names:
        validated_song["Data Match"] = "⚠️ Ingen komponist/writer i rapporten"
        return validated_song

    lookup = musicbrainz_lookup_work_writers(
        title=validated_song.get("Song title", ""),
        artist=validated_song.get("Artist", ""),
    )

    if lookup.get("status") == "error":
        validated_song["Data Match"] = f"⚠️ MusicBrainz-fejl: {lookup.get('message', '')}"
        return validated_song

    if lookup.get("status") == "not_found":
        validated_song["Data Match"] = "⚪ Ikke fundet i MusicBrainz"
        return validated_song

    candidates = lookup.get("candidates", []) or []

    if not candidates:
        validated_song["Data Match"] = "⚪ Ikke fundet i MusicBrainz"
        return validated_song

    # Vælg den kandidat, der matcher flest rapportnavne.
    best_candidate = None
    best_comparison = None
    best_score = -1

    for candidate in candidates:
        comparison = compare_report_names_to_musicbrainz(
            report_names=report_names,
            mb_names=candidate.get("writers", []),
        )

        score = len(comparison.get("matched", []))
        score += similarity(validated_song.get("Song title"), candidate.get("title"))

        if score > best_score:
            best_score = score
            best_candidate = candidate
            best_comparison = comparison

    if not best_candidate or not best_comparison:
        validated_song["Data Match"] = "⚪ MusicBrainz fandt værk, men ingen writer/composer-relationer"
        return validated_song

    mb_suggestions = best_comparison.get("musicbrainz_last_names", [])
    suggestion_text = "/".join(mb_suggestions) if mb_suggestions else "ingen forslag"

    if best_comparison.get("all_match"):
        validated_song["Data Match"] = f"✅ Match — MusicBrainz: {suggestion_text}"
    else:
        missing_text = "/".join(best_comparison.get("missing", [])) or "ukendt"
        validated_song["Data Match"] = (
            f"❌ Tjek: mangler {missing_text}. "
            f"MusicBrainz foreslår: {suggestion_text}"
        )

    return validated_song


def validate_songs_with_musicbrainz(
    songs: list[dict],
    validate_musicbrainz: bool,
    max_checks: int = 50,
    progress_callback=None,
    progress_start: float = 0,
    progress_end: float = 100,
    progress_label: str = "Validerer med MusicBrainz",
):
    """
    Validerer alle Existing music-sange mod MusicBrainz, hvis funktionen er slået til.

    max_checks er en sikkerhedsventil, så brugeren ikke utilsigtet laver hundredvis
    af API-kald på én gang. Sangene, der ikke bliver tjekket pga. grænsen, markeres
    tydeligt i previewet.
    """
    if not validate_musicbrainz:
        update_progress_safely(progress_callback, progress_end, "MusicBrainz-validering er slået fra")
        return songs, {
            "enabled": False,
            "checked": 0,
            "existing_music_total": sum(1 for song in songs if is_existing_music(song)),
            "warnings": [],
        }

    validated_songs = []
    checked = 0
    existing_music_total = sum(1 for song in songs if is_existing_music(song))
    warnings = []
    check_limit = min(existing_music_total, max_checks or existing_music_total) or 1

    update_progress_safely(
        progress_callback,
        progress_start,
        f"{progress_label}: starter ({existing_music_total} Existing music-tracks fundet)",
    )

    if musicbrainzngs is None:
        warnings.append("MusicBrainz-validering er slået til, men musicbrainzngs er ikke installeret.")

    for song in songs:
        if not is_existing_music(song):
            song_copy = dict(song)
            song_copy["Data Match"] = "— Ikke Existing music"
            validated_songs.append(song_copy)
            continue

        if max_checks and checked >= max_checks:
            song_copy = dict(song)
            song_copy["Data Match"] = "⏭️ Ikke tjekket — maksgrænse nået"
            validated_songs.append(song_copy)
            continue

        checked += 1
        validated_songs.append(validate_one_song_with_musicbrainz(song))

        # Opdater progressbaren efter hvert MusicBrainz-tjek.
        # Ved mange tracks kan brugeren derfor se, at appen stadig arbejder.
        fraction_done = min(checked / check_limit, 1)
        current_progress = progress_start + (progress_end - progress_start) * fraction_done
        update_progress_safely(
            progress_callback,
            current_progress,
            f"{progress_label}: {checked}/{check_limit} tjekket",
        )

        # Lille pause mellem nye API-kald. Cachede kald er hurtige, men ukendte
        # sange bør ikke spamme MusicBrainz unødigt.
        if checked < existing_music_total:
            time_module.sleep(1.0)

    if max_checks and existing_music_total > max_checks:
        warnings.append(
            f"MusicBrainz-validering tjekkede {checked} af {existing_music_total} Existing music-tracks. "
            "Hæv maksgrænsen i appen, hvis alle skal tjekkes."
        )

    update_progress_safely(progress_callback, progress_end, f"{progress_label}: færdig")

    return validated_songs, {
        "enabled": True,
        "checked": checked,
        "existing_music_total": existing_music_total,
        "warnings": warnings,
    }


def build_preview_rows(songs: list[dict], sheet_name: str = "", include_data_match: bool = False, max_rows: int = 500):
    """
    Bygger en enkel preview-tabel til Streamlit.

    Previewet påvirker ikke Excel-outputtet. Det er kun til kvalitetskontrol i appen.
    """
    preview_rows = []

    for song in songs[:max_rows]:
        row = {
            "Fane": sheet_name,
            "Song title": song.get("Song title", ""),
            "Artist": song.get("Artist", ""),
            "Composers": song.get("Composers", ""),
            "Writers": song.get("Writers", ""),
            "Min": song.get("Min", ""),
            "Sec": song.get("Sec", ""),
            "Track type": song.get("Track type", ""),
        }

        if include_data_match:
            row["Data Match"] = song.get("Data Match", "")

        preview_rows.append(row)

    return preview_rows


# ------------------------------------------------------------
# 9. UDFYLD SKABELONEN
# ------------------------------------------------------------

def find_template_music_columns(template_ws):
    """
    Finder kolonnerne i Music content-tabellen i skabelonen.

    Appen går ikke ud fra faste kolonnebogstaver.
    Den finder selv header-rækken og kolonnerne ud fra overskrifterne.
    """
    header_row, header_map = find_header_row(
        template_ws,
        required_headers=["Song title", "Artist", "Min", "Sec", "Track type"]
    )

    columns = {
        "Song title": get_col(header_map, ["Song title"]),
        "Artist": get_col(header_map, ["Artist"]),
        "Composers": get_col_contains(header_map, ["Composers"]),
        "Writers": get_col_contains(header_map, ["Writers"]),
        "Min": get_col(header_map, ["Min"]),
        "Sec": get_col(header_map, ["Sec"]),
        "Track type": get_col(header_map, ["Track type"]),
    }

    return header_row, columns


def write_metadata_to_template(template_ws, metadata: dict):
    """
    Skriver metadata ind i skabelonens øverste felter.

    Funktionen ændrer kun cellen til højre for label-feltet.
    """
    for target_label, value in metadata.items():
        target_cell = find_cell_right_of_label(template_ws, [target_label])

        if target_cell is None:
            raise ValueError(f"Kunne ikke finde metadatafeltet i skabelonen: {target_label}")

        # Vi skriver kun værdien. Formatering og layout bevares.
        target_cell.value = value


def copy_row_format(ws, source_row: int, target_row: int):
    """
    Kopierer kun formatering fra én række til en anden række.

    Det bruges, når en samlet rapport har flere tracks end de tomme rækker,
    skabelonen oprindeligt har. Vi ændrer ikke eksisterende rækker eller
    kolonner; vi fortsætter bare Music content-tabellen nedad og kopierer
    samme visuelle stil som skabelonens sidste musikrække.
    """
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col_index in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col_index)
        target_cell = ws.cell(row=target_row, column=col_index)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        # Vi kopierer ikke værdier. De nye celler skal være tomme,
        # indtil musikdata skrives ind.
        target_cell.value = None


def ensure_music_capacity(template_ws, first_data_row: int, needed_rows: int):
    """
    Sørger for, at Music content-tabellen har nok rækker til alle tracks.

    Skabelonen har normalt et fast antal tomme rækker. Ved en samlet rapport
    med alle faner kan der være flere tracks end det. Derfor udvider vi kun
    nederst i Music content-området og kopierer formateringen fra sidste
    eksisterende musikrække.
    """
    current_capacity = template_ws.max_row - first_data_row + 1

    if needed_rows <= current_capacity:
        return current_capacity

    source_format_row = template_ws.max_row
    final_row_needed = first_data_row + needed_rows - 1

    for row_number in range(template_ws.max_row + 1, final_row_needed + 1):
        copy_row_format(template_ws, source_format_row, row_number)

    return needed_rows


def write_music_to_template(template_ws, songs: list[dict], allow_expand: bool = False) -> dict:
    """
    Skriver musikdata ind i Music content-tabellen.

    Standardadfærd:
    - Ved én valgt fane og "én rapport pr. fane" skrives der kun i de
      eksisterende rækker i skabelonen.

    Samlet rapport:
    - Når allow_expand=True, udvider appen Music content-tabellen nedad, hvis
      der er flere tracks end de eksisterende tomme rækker. Den kopierer kun
      formateringen fra skabelonens sidste musikrække.
    """
    header_row, columns = find_template_music_columns(template_ws)
    first_data_row = header_row + 1

    warnings = []

    original_capacity = template_ws.max_row - first_data_row + 1

    if original_capacity <= 0:
        raise ValueError("Skabelonen har ingen tomme rækker under Music content-headeren.")

    if allow_expand:
        capacity = ensure_music_capacity(
            template_ws=template_ws,
            first_data_row=first_data_row,
            needed_rows=len(songs),
        )
        skipped_count = 0
        songs_to_write = songs

        if len(songs) > original_capacity:
            warnings.append(
                f"Music content-tabellen blev udvidet fra {original_capacity} til {len(songs)} rækker, "
                "så alle tracks fra alle faner kunne komme med i én samlet rapport."
            )
    else:
        capacity = original_capacity
        songs_to_write = songs[:capacity]
        skipped_count = max(0, len(songs) - capacity)

        if skipped_count > 0:
            warnings.append(
                f"Der var {len(songs)} musiknumre i kilden, men skabelonen har kun "
                f"plads til {capacity} eksisterende rækker. "
                f"{skipped_count} musiknumre blev derfor ikke skrevet ind. "
                f"Appen indsætter ikke nye rækker i denne eksportform."
            )

    # Ryd kun gamle værdier i de relevante musik-kolonner.
    # Dette ændrer ikke formatering, kolonnebredder eller layout.
    for row_number in range(first_data_row, first_data_row + capacity):
        for col_index in columns.values():
            template_ws.cell(row=row_number, column=col_index).value = None

    # Skriv sangene ind i skabelonen
    for index, song in enumerate(songs_to_write):
        row_number = first_data_row + index

        template_ws.cell(row=row_number, column=columns["Song title"]).value = song["Song title"]
        template_ws.cell(row=row_number, column=columns["Artist"]).value = song["Artist"]
        template_ws.cell(row=row_number, column=columns["Composers"]).value = song["Composers"]
        template_ws.cell(row=row_number, column=columns["Writers"]).value = song["Writers"]
        template_ws.cell(row=row_number, column=columns["Min"]).value = song["Min"]
        template_ws.cell(row=row_number, column=columns["Sec"]).value = song["Sec"]
        template_ws.cell(row=row_number, column=columns["Track type"]).value = song["Track type"]

    return {
        "written_count": len(songs_to_write),
        "skipped_count": skipped_count,
        "capacity": capacity,
        "warnings": warnings,
    }


# ------------------------------------------------------------
# 10. HOVEDFUNKTION: FRA TO UPLOADEDE FILER TIL FÆRDIG EXCEL
# ------------------------------------------------------------

def get_sheet_names(source_file_bytes: bytes) -> list[str]:
    """
    Læser alle fanenavne fra kildefilen.

    Dette bruges til den dynamiske dropdown-menu.
    """
    workbook = load_workbook(
        io.BytesIO(source_file_bytes),
        read_only=True,
        data_only=True,
    )
    return workbook.sheetnames


def update_progress_safely(progress_callback, percent: float, message: str):
    """
    Opdaterer progress baren uden at risikere, at selve databehandlingen crasher.

    Funktionen får en callback fra Streamlit-UI'et. Hvis callbacken mod forventning
    fejler, ignorerer vi fejlen, fordi rapporten stadig skal kunne laves færdig.
    """
    if progress_callback is None:
        return

    try:
        clean_percent = int(max(0, min(100, round(percent))))
        progress_callback(clean_percent, message)
    except Exception:
        # Progressbaren er kun brugerflade. Den må aldrig stoppe rapportgenereringen.
        pass


def process_files(
    template_file_bytes: bytes,
    source_file_bytes: bytes,
    selected_sheet: str,
    selected_country: str,
    production_type: str,
    auto_group_special: bool = True,
    validate_musicbrainz: bool = False,
    max_musicbrainz_checks: int = 50,
    progress_callback=None,
):
    """
    Hele databehandlingen samlet ét sted.

    Input:
    - template_file_bytes: bytes fra uploadet tv-appl-en.xlsx
    - source_file_bytes: bytes fra uploadet musikrapport-export
    - selected_sheet: fanen brugeren har valgt i dropdown

    Output:
    - output_bytes: færdig Excel-fil som bytes
    - summary: info til brugerfladen
    """
    update_progress_safely(progress_callback, 3, "Åbner kildefilen")

    # Åbn kildefilen
    source_wb = load_workbook(
        io.BytesIO(source_file_bytes),
        read_only=False,
        data_only=True,
    )

    if selected_sheet not in source_wb.sheetnames:
        raise ValueError(f"Den valgte fane findes ikke i kildefilen: {selected_sheet}")

    source_ws = source_wb[selected_sheet]
    update_progress_safely(progress_callback, 12, f"Læser fanen: {selected_sheet}")

    # Udtræk metadata og musikdata
    metadata = extract_metadata(source_ws)
    songs = extract_music_rows(source_ws)
    update_progress_safely(progress_callback, 28, f"Fandt {len(songs)} musiklinjer i {selected_sheet}")

    original_song_count = len(songs)

    # Brug land + produktionstype til at anvende de relevante rapportregler.
    songs, rule_summary = apply_country_production_rules(
        songs=songs,
        selected_country=selected_country,
        production_type=production_type,
        auto_group_special=auto_group_special,
    )
    update_progress_safely(progress_callback, 38, "Anvender land- og produktionstype-regler")

    # Valgfri MusicBrainz-validering. Den påvirker kun Streamlit-previewet,
    # ikke selve Excel-skabelonen.
    songs, musicbrainz_summary = validate_songs_with_musicbrainz(
        songs=songs,
        validate_musicbrainz=validate_musicbrainz,
        max_checks=max_musicbrainz_checks,
        progress_callback=progress_callback,
        progress_start=40,
        progress_end=75,
        progress_label="Validerer komponister med MusicBrainz",
    )

    update_progress_safely(progress_callback, 78, "Åbner Excel-skabelonen")

    # Åbn skabelonen
    template_wb = load_workbook(
        io.BytesIO(template_file_bytes),
        read_only=False,
        data_only=False,
    )

    # Brug Application-fanen, hvis den findes. Ellers brug aktiv fane.
    if "Application" in template_wb.sheetnames:
        template_ws = template_wb["Application"]
    else:
        template_ws = template_wb.active

    # Skriv data ind i skabelonen
    write_metadata_to_template(template_ws, metadata)
    write_report_context_to_template(template_ws, selected_country, production_type)
    write_result = write_music_to_template(template_ws, songs)
    update_progress_safely(progress_callback, 90, "Skriver musikdata ind i skabelonen")

    # Læg regelmotorens advarsler sammen med Excel-skrivningens advarsler.
    write_result["warnings"] = (
        rule_summary.get("warnings", [])
        + musicbrainz_summary.get("warnings", [])
        + write_result.get("warnings", [])
    )

    # Gem workbook i hukommelsen i stedet for på disk.
    # Det gør, at Streamlit kan sende filen direkte til download-knappen.
    update_progress_safely(progress_callback, 96, "Gemmer den færdige Excel-fil")
    output = io.BytesIO()
    template_wb.save(output)
    output.seek(0)
    update_progress_safely(progress_callback, 100, "Færdig")

    summary = {
        "selected_sheet": selected_sheet,
        "metadata": metadata,
        "country": selected_country,
        "production_type": production_type,
        "rule_summary": rule_summary,
        "musicbrainz_summary": musicbrainz_summary,
        "preview_rows": build_preview_rows(
            songs,
            sheet_name=selected_sheet,
            include_data_match=validate_musicbrainz,
        ),
        "song_count_found": original_song_count,
        "song_count_after_rules": len(songs),
        "song_count_written": write_result["written_count"],
        "song_count_skipped": write_result["skipped_count"],
        "template_capacity": write_result["capacity"],
        "warnings": write_result["warnings"],
    }

    return output.getvalue(), summary


def make_unique_zip_name(filename: str, used_names: set[str]) -> str:
    """
    Sørger for, at filnavne inde i ZIP-filen er unikke.

    Hvis to faner mod forventning giver samme filnavn, laver vi fx:
    - rapport.xlsx
    - rapport_2.xlsx
    """
    if filename not in used_names:
        used_names.add(filename)
        return filename

    if "." in filename:
        base, extension = filename.rsplit(".", 1)
        extension = "." + extension
    else:
        base, extension = filename, ""

    counter = 2
    while True:
        candidate = f"{base}_{counter}{extension}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1


def write_optional_metadata_value(template_ws, target_label: str, value):
    """
    Skriver et valgfrit metadatafelt, hvis feltet findes i skabelonen.

    Bruges fx til "Total number of episodes" i en samlet rapport.
    Hvis feltet ikke findes, springer appen bare videre uden at crashe.
    """
    target_cell = find_cell_right_of_label(template_ws, [target_label])

    if target_cell is not None:
        target_cell.value = value


def process_all_sheets_combined(
    template_file_bytes: bytes,
    source_file_bytes: bytes,
    sheet_names: list[str],
    selected_country: str,
    production_type: str,
    auto_group_special: bool = True,
    validate_musicbrainz: bool = False,
    max_musicbrainz_checks: int = 50,
    progress_callback=None,
):
    """
    Behandler alle faner og samler ALLE tracks i ÉN musikrapport.

    Det betyder:
    - Outputtet er én Excel-fil.
    - Outputtet har stadig kun én Application-rapportfane fra skabelonen.
    - Alle tracks fra alle valgte kildefaner skrives samlet ned i samme
      Music content-tabel.

    Hvis der er flere tracks end skabelonens oprindelige tomme rækker, udvider
    appen Music content-området nedad og kopierer formateringen fra sidste
    eksisterende musikrække. Det er nødvendigt for at få alle tracks med i én
    samlet rapport.
    """
    update_progress_safely(progress_callback, 3, "Åbner kildefilen")

    # Åbn kildefilen én gang
    source_wb = load_workbook(
        io.BytesIO(source_file_bytes),
        read_only=False,
        data_only=True,
    )

    all_songs = []
    summaries = []
    errors = []
    first_metadata = None

    total_sheets = max(len(sheet_names), 1)

    for sheet_index, sheet_name in enumerate(sheet_names, start=1):
        try:
            update_progress_safely(
                progress_callback,
                5 + (35 * (sheet_index - 1) / total_sheets),
                f"Læser fane {sheet_index}/{total_sheets}: {sheet_name}",
            )
            if sheet_name not in source_wb.sheetnames:
                raise ValueError(f"Fane findes ikke i kildefilen: {sheet_name}")

            source_ws = source_wb[sheet_name]

            # Udtræk data fra kildefanen
            metadata = extract_metadata(source_ws)
            songs = extract_music_rows(source_ws)

            # Gem metadata fra første fane, der kan behandles.
            # I samlede rapporter er Production company/title/broadcaster normalt ens
            # på tværs af faner. Episode number ændrer vi længere nede.
            if first_metadata is None:
                first_metadata = metadata

            all_songs.extend(songs)

            update_progress_safely(
                progress_callback,
                5 + (35 * sheet_index / total_sheets),
                f"Fane {sheet_index}/{total_sheets} læst: {len(songs)} tracks",
            )

            summaries.append(
                {
                    "selected_sheet": sheet_name,
                    "metadata": metadata,
                    "song_count_found": len(songs),
                    "song_count_written": len(songs),
                    "song_count_skipped": 0,
                    "warnings": [],
                }
            )

        except Exception as error:
            # Hvis én fane fejler, skal hele appen ikke dø.
            # Vi gemmer fejlen og fortsætter med næste fane.
            errors.append(
                {
                    "sheet": sheet_name,
                    "error": str(error),
                }
            )

    if not summaries:
        error_text = "; ".join(
            f"{item['sheet']}: {item['error']}" for item in errors[:5]
        )
        raise ValueError(
            "Ingen faner kunne behandles. "
            f"De første fejl var: {error_text}"
        )

    original_all_song_count = len(all_songs)
    update_progress_safely(progress_callback, 45, f"Samler {len(all_songs)} tracks fra alle faner")

    # Når alle tracks er samlet fra alle faner, kan vi samle gentagne bumpers/vignetter
    # på tværs af hele uploadet. Det er især nyttigt for fx norske bumpers.
    all_songs, rule_summary = apply_country_production_rules(
        songs=all_songs,
        selected_country=selected_country,
        production_type=production_type,
        auto_group_special=auto_group_special,
    )
    update_progress_safely(progress_callback, 55, "Anvender land- og produktionstype-regler på samlet rapport")

    # Valgfri MusicBrainz-validering på den samlede trackliste.
    all_songs, musicbrainz_summary = validate_songs_with_musicbrainz(
        songs=all_songs,
        validate_musicbrainz=validate_musicbrainz,
        max_checks=max_musicbrainz_checks,
        progress_callback=progress_callback,
        progress_start=58,
        progress_end=82,
        progress_label="Validerer samlet rapport med MusicBrainz",
    )

    update_progress_safely(progress_callback, 85, "Åbner Excel-skabelonen")

    # Åbn skabelonen
    template_wb = load_workbook(
        io.BytesIO(template_file_bytes),
        read_only=False,
        data_only=False,
    )

    # Brug Application-fanen, hvis den findes. Ellers brug aktiv fane.
    if "Application" in template_wb.sheetnames:
        template_ws = template_wb["Application"]
    else:
        template_ws = template_wb.active

    # Metadata til én samlet rapport:
    # Vi bruger metadata fra første behandlede fane, men gør Episode number tydelig,
    # så rapporten ikke ligner én enkelt episode.
    combined_metadata = dict(first_metadata or {})
    combined_metadata["Episode number"] = "All uploaded sheets"

    write_metadata_to_template(template_ws, combined_metadata)
    write_report_context_to_template(template_ws, selected_country, production_type)
    write_optional_metadata_value(template_ws, "Total number of episodes", len(summaries))

    # Skriv ALLE tracks ind i den samme Music content-tabel.
    write_result = write_music_to_template(template_ws, all_songs, allow_expand=True)
    update_progress_safely(progress_callback, 93, "Skriver alle tracks ind i samlet rapport")

    # Tilføj eventuelle warnings fra regelmotoren og selve skrivningen til summary.
    write_result["warnings"] = (
        rule_summary.get("warnings", [])
        + musicbrainz_summary.get("warnings", [])
        + write_result.get("warnings", [])
    )
    for warning in write_result["warnings"]:
        summaries[0].setdefault("warnings", []).append(warning)

    update_progress_safely(progress_callback, 97, "Gemmer samlet Excel-fil")
    output = io.BytesIO()
    template_wb.save(output)
    output.seek(0)
    update_progress_safely(progress_callback, 100, "Færdig")

    summary = {
        "mode": "combined",
        "processed_count": len(summaries),
        "error_count": len(errors),
        "summaries": summaries,
        "errors": errors,
        "metadata": combined_metadata,
        "country": selected_country,
        "production_type": production_type,
        "rule_summary": rule_summary,
        "musicbrainz_summary": musicbrainz_summary,
        "preview_rows": build_preview_rows(
            all_songs,
            sheet_name="Alle faner",
            include_data_match=validate_musicbrainz,
        ),
        "song_count_found": original_all_song_count,
        "song_count_after_rules": len(all_songs),
        "song_count_written": write_result["written_count"],
        "song_count_skipped": write_result["skipped_count"],
        "template_capacity": write_result["capacity"],
        "warnings": write_result["warnings"],
    }

    return output.getvalue(), summary

def process_all_sheets_separate_files(
    template_file_bytes: bytes,
    source_file_bytes: bytes,
    sheet_names: list[str],
    selected_country: str,
    production_type: str,
    auto_group_special: bool = True,
    validate_musicbrainz: bool = False,
    max_musicbrainz_checks: int = 50,
    progress_callback=None,
):
    """
    Behandler alle faner og laver én separat Excel-rapport pr. fane.

    Outputtet pakkes i en ZIP-fil, så brugeren kan downloade det hele på én gang.
    Dette er den mest klassiske "én musikrapport pr. fil"-løsning.
    """
    update_progress_safely(progress_callback, 2, "Starter ZIP-eksport med én rapport pr. fane")

    zip_buffer = io.BytesIO()
    summaries = []
    errors = []
    used_zip_names = set()

    total_sheets = max(len(sheet_names), 1)

    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        for sheet_index, sheet_name in enumerate(sheet_names, start=1):
            try:
                sheet_start = 5 + (90 * (sheet_index - 1) / total_sheets)
                sheet_end = 5 + (90 * sheet_index / total_sheets)

                def sheet_progress(local_percent, local_message, sheet_name=sheet_name, sheet_index=sheet_index, sheet_start=sheet_start, sheet_end=sheet_end):
                    overall_percent = sheet_start + (sheet_end - sheet_start) * (local_percent / 100)
                    update_progress_safely(
                        progress_callback,
                        overall_percent,
                        f"Fane {sheet_index}/{total_sheets} — {sheet_name}: {local_message}",
                    )

                finished_file_bytes, summary = process_files(
                    template_file_bytes=template_file_bytes,
                    source_file_bytes=source_file_bytes,
                    selected_sheet=sheet_name,
                    selected_country=selected_country,
                    production_type=production_type,
                    auto_group_special=auto_group_special,
                    validate_musicbrainz=validate_musicbrainz,
                    max_musicbrainz_checks=max_musicbrainz_checks,
                    progress_callback=sheet_progress,
                )

                output_filename = f"tv-appl-en_udfyldt_{safe_filename(sheet_name)}.xlsx"
                output_filename = make_unique_zip_name(output_filename, used_zip_names)

                zip_file.writestr(output_filename, finished_file_bytes)
                summary["output_filename"] = output_filename
                summaries.append(summary)
                update_progress_safely(
                    progress_callback,
                    sheet_end,
                    f"Fane {sheet_index}/{total_sheets} færdig: {sheet_name}",
                )

            except Exception as error:
                # Hvis én fane fejler, skal hele appen ikke dø.
                # Vi gemmer fejlen og fortsætter med næste fane.
                errors.append(
                    {
                        "sheet": sheet_name,
                        "error": str(error),
                    }
                )

    if not summaries:
        error_text = "; ".join(
            f"{item['sheet']}: {item['error']}" for item in errors[:5]
        )
        raise ValueError(
            "Ingen faner kunne behandles. "
            f"De første fejl var: {error_text}"
        )

    update_progress_safely(progress_callback, 97, "Samler ZIP-filen")
    zip_buffer.seek(0)

    total_aggregated_original_rows = sum(
        item.get("rule_summary", {}).get("aggregated_original_rows", 0)
        for item in summaries
    )
    total_aggregated_groups = sum(
        item.get("rule_summary", {}).get("aggregated_groups", 0)
        for item in summaries
    )
    total_collapsed_lines = sum(
        item.get("rule_summary", {}).get("collapsed_lines", 0)
        for item in summaries
    )

    preview_rows = []
    for item in summaries:
        preview_rows.extend(item.get("preview_rows", []))

    total_musicbrainz_checked = sum(
        item.get("musicbrainz_summary", {}).get("checked", 0)
        for item in summaries
    )
    total_existing_music = sum(
        item.get("musicbrainz_summary", {}).get("existing_music_total", 0)
        for item in summaries
    )
    musicbrainz_warnings = []
    for item in summaries:
        musicbrainz_warnings.extend(item.get("musicbrainz_summary", {}).get("warnings", []))

    summary = {
        "mode": "separate_files",
        "processed_count": len(summaries),
        "error_count": len(errors),
        "summaries": summaries,
        "errors": errors,
        "country": selected_country,
        "production_type": production_type,
        "rule_summary": {
            "aggregated_original_rows": total_aggregated_original_rows,
            "aggregated_groups": total_aggregated_groups,
            "collapsed_lines": total_collapsed_lines,
            "country": selected_country,
            "production_type": production_type,
            "auto_group_special": auto_group_special,
        },
        "musicbrainz_summary": {
            "enabled": validate_musicbrainz,
            "checked": total_musicbrainz_checked,
            "existing_music_total": total_existing_music,
            "warnings": musicbrainz_warnings,
        },
        "preview_rows": preview_rows[:500],
        "song_count_found": sum(item.get("song_count_found", 0) for item in summaries),
        "song_count_after_rules": sum(item.get("song_count_after_rules", item.get("song_count_found", 0)) for item in summaries),
        "song_count_written": sum(item.get("song_count_written", 0) for item in summaries),
        "song_count_skipped": sum(item.get("song_count_skipped", 0) for item in summaries),
    }

    update_progress_safely(progress_callback, 100, "Færdig")

    return zip_buffer.getvalue(), summary


def safe_filename(text: str) -> str:
    """
    Gør et fanenavn sikkert som filnavn.
    """
    text = safe_text(text)
    text = re.sub(r"[^A-Za-z0-9ÆØÅæøå._ -]+", "_", text)
    text = re.sub(r"\s+", "_", text).strip("_")
    return text or "rapport"


# ------------------------------------------------------------
# 11. STREAMLIT UI: UPLOAD, DROPDOWN, KNAP OG DOWNLOAD
# ------------------------------------------------------------

st.divider()

col1, col2 = st.columns(2)

with col1:
    template_file = st.file_uploader(
        "1. Upload skabelon",
        type=["xlsx"],
        help="Upload tv-appl-en.xlsx",
        key="template_uploader",
    )

with col2:
    source_file = st.file_uploader(
        "2. Upload kildefil",
        type=["xlsx"],
        help="Upload musikrapport-exporten med mange faner",
        key="source_uploader",
    )


st.subheader("3. Land og produktionstype")
selected_country = st.selectbox(
    "Vælg land / NCB-prisliste",
    options=get_country_names(),
    index=get_country_names().index("Norge") if "Norge" in get_country_names() else 0,
    help="Landet bestemmer hvilke rapporteringsregler og produktionstyper, der vises.",
)

selected_production_type = st.selectbox(
    "Vælg produktionstype",
    options=get_production_types_for_country(selected_country),
    index=0,
    help="Produktionstypen bruges til at vælge de rigtige automatiske rapportregler.",
)

auto_group_special = st.checkbox(
    "Saml automatisk ens bumpers/vignetter",
    value=True,
    help=(
        "Når denne er slået til, samler appen gentagne bumpers/vignetter til én linje "
        "med fx 'x 12'. Det er især relevant for norske rapporter."
    ),
)

st.info(COUNTRY_CONFIG[selected_country]["rule_note"])

validate_musicbrainz = st.toggle(
    "Validér komponister i Existing music med MusicBrainz",
    value=False,
    help=(
        "Når denne er slået til, slår appen Existing music-tracks op i MusicBrainz "
        "og viser en Data Match-kolonne i previewet. Det kan tage lidt tid, fordi "
        "der laves API-kald for hvert tjekket track."
    ),
)

max_musicbrainz_checks = 50
if validate_musicbrainz:
    st.warning(
        "MusicBrainz-validering er et hjælpetjek — ikke en juridisk facitliste. "
        "Obskur library music og meget nye værker findes ofte ikke i databasen."
    )
    max_musicbrainz_checks = st.number_input(
        "Maks antal Existing music-tracks, der må tjekkes mod MusicBrainz",
        min_value=1,
        max_value=500,
        value=50,
        step=10,
        help="Sæt tallet højere, hvis alle Existing music-tracks skal valideres. Lavere tal gør appen hurtigere.",
    )

selected_sheet = None
process_mode = None
export_mode = None
sheet_names = []

if source_file is not None:
    try:
        source_bytes_for_sheets = source_file.getvalue()
        sheet_names = get_sheet_names(source_bytes_for_sheets)

        if sheet_names:
            st.caption(f"Fandt {len(sheet_names)} faner i kildefilen.")

            process_mode = st.radio(
                "4. Hvad vil du udtrække?",
                options=["Én valgt fane", "Alle faner"],
                horizontal=True,
            )

            if process_mode == "Én valgt fane":
                selected_sheet = st.selectbox(
                    "Vælg hvilken fane der skal udtrækkes fra",
                    options=sheet_names,
                    index=0,
                )
            else:
                export_mode = st.radio(
                    "5. Hvordan vil du eksportere alle faner?",
                    options=["Samlet musikrapport", "Én musikrapport pr. fane"],
                    horizontal=False,
                    help=(
                        "Samlet musikrapport giver én Excel-fil, hvor alle tracks fra alle faner "
                        "samles i den samme Music content-tabel. "
                        "Én musikrapport pr. fane giver en ZIP-fil med separate Excel-filer."
                    ),
                )

                if export_mode == "Samlet musikrapport":
                    st.info(
                        f"Alle {len(sheet_names)} faner bliver behandlet. "
                        "Du får én samlet Excel-fil, hvor alle tracks ligger i den samme Music content-tabel."
                    )
                else:
                    st.info(
                        f"Alle {len(sheet_names)} faner bliver behandlet. "
                        "Du får én ZIP-fil med én separat Excel-rapport pr. kildefane."
                    )

                with st.expander("Se faner, der kommer med"):
                    for sheet_name in sheet_names:
                        st.write(f"- {sheet_name}")
        else:
            st.warning("Kildefilen indeholder ingen faner.")

    except InvalidFileException:
        st.error("Kildefilen ser ikke ud til at være en gyldig .xlsx-fil.")

    except Exception:
        st.error("Kunne ikke læse fanerne i kildefilen.")
        with st.expander("Vis tekniske detaljer"):
            st.code(traceback.format_exc())


st.divider()

button_disabled = (
    template_file is None
    or source_file is None
    or not sheet_names
    or selected_country is None
    or selected_production_type is None
    or process_mode is None
    or (process_mode == "Én valgt fane" and selected_sheet is None)
    or (process_mode == "Alle faner" and export_mode is None)
)

if st.button("🚀 Udfyld Rapport", type="primary", disabled=button_disabled):
    try:
        # Progressbar, så brugeren kan følge med i lange kørsler
        # — især når alle faner eller MusicBrainz-validering er slået til.
        progress_bar = st.progress(0, text="Starter...")
        progress_caption = st.empty()

        def ui_progress(percent, message):
            clean_percent = int(max(0, min(100, percent)))
            progress_bar.progress(clean_percent, text=message)
            progress_caption.caption(f"{clean_percent}% — {message}")

        with st.spinner("Udfylder rapporten..."):
            ui_progress(1, "Læser uploadede filer")
            template_bytes = template_file.getvalue()
            source_bytes = source_file.getvalue()

            if process_mode == "Alle faner" and export_mode == "Samlet musikrapport":
                finished_file_bytes, summary = process_all_sheets_combined(
                    template_file_bytes=template_bytes,
                    source_file_bytes=source_bytes,
                    sheet_names=sheet_names,
                    selected_country=selected_country,
                    production_type=selected_production_type,
                    auto_group_special=auto_group_special,
                    validate_musicbrainz=validate_musicbrainz,
                    max_musicbrainz_checks=max_musicbrainz_checks,
                    progress_callback=ui_progress,
                )
                output_filename = "tv-appl-en_samlet_musikrapport_alle_tracks.xlsx"
                download_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                download_label = "⬇️ Download samlet musikrapport med alle tracks"

            elif process_mode == "Alle faner" and export_mode == "Én musikrapport pr. fane":
                finished_file_bytes, summary = process_all_sheets_separate_files(
                    template_file_bytes=template_bytes,
                    source_file_bytes=source_bytes,
                    sheet_names=sheet_names,
                    selected_country=selected_country,
                    production_type=selected_production_type,
                    auto_group_special=auto_group_special,
                    validate_musicbrainz=validate_musicbrainz,
                    max_musicbrainz_checks=max_musicbrainz_checks,
                    progress_callback=ui_progress,
                )
                output_filename = "tv-appl-en_rapporter_pr_fane.zip"
                download_mime = "application/zip"
                download_label = "⬇️ Download ZIP med én rapport pr. fane"

            else:
                finished_file_bytes, summary = process_files(
                    template_file_bytes=template_bytes,
                    source_file_bytes=source_bytes,
                    selected_sheet=selected_sheet,
                    selected_country=selected_country,
                    production_type=selected_production_type,
                    auto_group_special=auto_group_special,
                    validate_musicbrainz=validate_musicbrainz,
                    max_musicbrainz_checks=max_musicbrainz_checks,
                    progress_callback=ui_progress,
                )
                summary["mode"] = "single"
                output_filename = f"tv-appl-en_udfyldt_{safe_filename(selected_sheet)}.xlsx"
                download_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                download_label = "⬇️ Download udfyldt Excel-fil"

            # Gem resultatet i session_state, så download-knappen stadig virker
            # efter Streamlit har genindlæst siden.
            st.session_state["finished_file_bytes"] = finished_file_bytes
            st.session_state["output_filename"] = output_filename
            st.session_state["download_mime"] = download_mime
            st.session_state["download_label"] = download_label
            st.session_state["summary"] = summary
            ui_progress(100, "Rapporten er færdig og klar til download")

        st.success("Rapporten er udfyldt og klar til download.")

    except InvalidFileException:
        st.error("En af filerne ser ikke ud til at være en gyldig .xlsx-fil.")

    except Exception:
        st.error("Der skete en fejl under udfyldningen af rapporten.")
        st.write("Tjek at skabelonen og kildefilen har den forventede struktur.")

        with st.expander("Vis tekniske detaljer"):
            st.code(traceback.format_exc())


# ------------------------------------------------------------
# 12. VIS RESULTAT OG DOWNLOAD-KNAP
# ------------------------------------------------------------

if "finished_file_bytes" in st.session_state:
    summary = st.session_state.get("summary", {})

    st.subheader("✅ Klar til download")

    if summary.get("mode") in ["combined", "separate_files"]:
        st.write("**Udtræk:** Alle faner")
        st.write(f"**Land:** {summary.get('country', '')}")
        st.write(f"**Produktionstype:** {summary.get('production_type', '')}")

        if summary.get("mode") == "combined":
            st.write("**Eksport:** Samlet musikrapport — én Excel-fil med alle tracks samlet i samme Music content-tabel")
        else:
            st.write("**Eksport:** Én musikrapport pr. fane — ZIP-fil med separate Excel-filer")

        st.write(f"**Faner behandlet:** {summary.get('processed_count', 0)}")
        st.write(f"**Faner med fejl:** {summary.get('error_count', 0)}")
        st.write(f"**Musiknumre fundet i alt:** {summary.get('song_count_found', 0)}")
        st.write(f"**Musiknumre skrevet i alt:** {summary.get('song_count_written', 0)}")

        if summary.get("song_count_skipped", 0) > 0:
            st.write(f"**Musiknumre sprunget over i alt:** {summary.get('song_count_skipped', 0)}")

        if summary.get("mode") == "combined":
            with st.expander("Se hvor mange tracks der blev hentet fra hver fane"):
                for item in summary.get("summaries", []):
                    st.write(
                        f"- **{item.get('selected_sheet', '')}**: "
                        f"{item.get('song_count_found', 0)} tracks"
                    )
        else:
            with st.expander("Se separate rapportfiler"):
                for item in summary.get("summaries", []):
                    st.write(
                        f"- **{item.get('selected_sheet', '')}** → "
                        f"{item.get('output_filename', '')} "
                        f"({item.get('song_count_written', 0)} musiknumre skrevet)"
                    )

        if summary.get("errors"):
            st.warning(
                "Nogle faner kunne ikke behandles. "
                "De fungerende faner er stadig med i downloadfilen."
            )
            with st.expander("Se fejl pr. fane"):
                for item in summary.get("errors", []):
                    st.write(f"- **{item.get('sheet', '')}**: {item.get('error', '')}")

        all_warnings = []
        for item in summary.get("summaries", []):
            for warning in item.get("warnings", []):
                all_warnings.append(f"{item.get('selected_sheet', '')}: {warning}")

        for warning in all_warnings:
            st.warning(warning)

    else:
        st.write(f"**Valgt fane:** {summary.get('selected_sheet', '')}")
        st.write(f"**Land:** {summary.get('country', '')}")
        st.write(f"**Produktionstype:** {summary.get('production_type', '')}")
        st.write(f"**Musiknumre fundet i kildefilen:** {summary.get('song_count_found', 0)}")
        st.write(f"**Musiknumre skrevet i skabelonen:** {summary.get('song_count_written', 0)}")

        if summary.get("song_count_skipped", 0) > 0:
            st.write(f"**Musiknumre sprunget over:** {summary.get('song_count_skipped', 0)}")

        warnings = summary.get("warnings", [])
        for warning in warnings:
            st.warning(warning)

        metadata = summary.get("metadata", {})
        with st.expander("Se metadata, der blev overført"):
            st.write(metadata)

    rule_summary = summary.get("rule_summary", {})
    if rule_summary and rule_summary.get("collapsed_lines", 0) > 0:
        st.success(
            f"Smart-regel brugt: {rule_summary.get('aggregated_original_rows', 0)} "
            f"bumper-/vignette-linjer blev samlet til "
            f"{rule_summary.get('aggregated_groups', 0)} linjer."
        )

    musicbrainz_summary = summary.get("musicbrainz_summary", {})
    if musicbrainz_summary and musicbrainz_summary.get("enabled"):
        st.info(
            f"MusicBrainz-validering: {musicbrainz_summary.get('checked', 0)} "
            f"af {musicbrainz_summary.get('existing_music_total', 0)} Existing music-tracks blev tjekket."
        )
        for warning in musicbrainz_summary.get("warnings", []):
            st.warning(warning)

    preview_rows = summary.get("preview_rows", [])
    if preview_rows:
        st.subheader("Preview af musikdata")
        st.dataframe(
            preview_rows,
            hide_index=True,
            use_container_width=True,
        )
        if len(preview_rows) >= 500:
            st.caption("Previewet viser de første 500 linjer for at holde appen hurtig.")

    st.download_button(
        label=st.session_state.get("download_label", "⬇️ Download fil"),
        data=st.session_state["finished_file_bytes"],
        file_name=st.session_state["output_filename"],
        mime=st.session_state.get("download_mime", "application/octet-stream"),
        type="primary",
    )


# ------------------------------------------------------------
# 13. LILLE HJÆLPETEKST TIL BRUGEREN
# ------------------------------------------------------------

with st.expander("Hvad gør appen helt præcist?"):
    st.markdown(
        """
Appen gør følgende:

1. Læser alle faner i kildefilen.
2. Lader dig vælge land og produktionstype.
3. Lader dig vælge enten én fane eller alle faner.
4. Finder metadatafelterne:
   - Production Company
   - Production Title
   - Episode Number
   - Network/Station
5. Finder tabellen med musikdata via kolonnen **Seq #**.
6. Grupperer linjerne under hver sang, så komponister/forfattere tilhører den rigtige sang.
7. Udtrækker efternavne og fjerner foreningskoder som `[PRS]`, `[BMI]` og `(BMI)`.
8. Splitter **Music Duration** til minutter og sekunder.
9. Oversætter **Music Source** til skabelonens track types.
10. Samler automatisk gentagne bumpers/vignetter, hvis land/produktionstype kræver det.
11. Kan valgfrit slå Existing music-tracks op i MusicBrainz og vise en **Data Match**-kolonne i previewet.
12. Skriver værdierne ind i de eksisterende celler i skabelonen.
13. Giver dig enten én færdig Excel-fil, én samlet musikrapport med alle tracks i samme tabel eller en ZIP-fil med én Excel-rapport pr. fane direkte i browseren.
"""
    )
